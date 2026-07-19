import os
import shutil
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from scipy.optimize import curve_fit
from scipy.stats import ttest_rel, ttest_ind


plt.rcParams.update({
    "axes.labelsize": 20,    # GP,T,P/L
    "axes.titlesize": 18,
    "xtick.labelsize": 16,   # numeros eje x
    "ytick.labelsize": 16,   # numeros eje y
    "legend.fontsize": 15,
    "font.size": 16,
})

VOLUMEN_A_PL = {
    "0ul": 0.0,
    "2ul": 0.04,
    "8ul": 0.16,
    "16ul": 0.32,
}

COLOR_PL = {
    0.0:  "tab:blue",
    0.04: "tab:orange",
    0.16: "tab:green",
    0.32: "tab:red",
}

CEPAS = ["144", "145"]

PEPTIDOS_PIAM = {
    "piam217": "piam21-7",
    "piam234": "piam23-4",
    "piam297": "piam29-7",
}

PEPTIDOS_LL37 = {
    "LL-37": "LL-37",
}

PEPTIDOS = {**PEPTIDOS_PIAM, **PEPTIDOS_LL37}


MARKER_CEPA = {
    "144": "^",
    "145": "o",
}

COLOR_CEPA = {
    "144": "tab:blue",
    "145": "tab:red",
}


MARGEN_YLIM = 0.05


TIEMPO_CON_PEPTIDO_MIN = 20


def cargar_excel(ruta):
   
    return openpyxl.load_workbook(ruta, data_only=True)


def normalizar_nombre_hoja(nombre):
   
    return nombre.lower().replace(" ", "")


def buscar_hoja_piam(wb, cepa, peptido_corto, volumen):
    
    if volumen == "0ul":
        patron = f"sa{cepa}-0ul"
    else:
        patron = f"sa{cepa}-{peptido_corto}-{volumen}"

    patron_norm = normalizar_nombre_hoja(patron)
    for nombre in wb.sheetnames:
        if normalizar_nombre_hoja(nombre) == patron_norm:
            return wb[nombre]
    return None


def buscar_hoja_ll37(wb, cepa, volumen):
    
    if volumen == "0ul":
        patron_norm = normalizar_nombre_hoja(f"sa{cepa}-0ul")
        for nombre in wb.sheetnames:
            if normalizar_nombre_hoja(nombre) == patron_norm:
                return wb[nombre]
        return None

    patron_base = normalizar_nombre_hoja(f"sa{cepa}-LL-37-{volumen}")

    for nombre in wb.sheetnames:
        norm = normalizar_nombre_hoja(nombre)
        if norm == patron_base or norm.startswith(patron_base):
            return wb[nombre]
    return None


def obtener_hoja(wb, cepa, peptido_corto, volumen, es_ll37):
    
    if volumen == "0ul":
        return buscar_hoja_piam(wb, cepa, "", "0ul")
    if es_ll37:
        return buscar_hoja_ll37(wb, cepa, volumen)
    return buscar_hoja_piam(wb, cepa, peptido_corto, volumen)


def leer_bloque_gp(ws):
    temperaturas = []
    promedios = []
    desviaciones = []

    fila = 3
    while fila <= ws.max_row:
        temp = ws.cell(row=fila, column=1).value
        if not isinstance(temp, (int, float)):
            break
        gp_prom = ws.cell(row=fila, column=6).value
        desv = ws.cell(row=fila, column=7).value

        if isinstance(gp_prom, (int, float)) and isinstance(desv, (int, float)):
            temperaturas.append(temp)
            promedios.append(gp_prom)
            desviaciones.append(desv)
        fila += 1

    return temperaturas, promedios, desviaciones


def leer_bloque_intensidad(ws, tipo):
    fila_titulo = None
    for fila in range(1, ws.max_row + 1):
        val = ws.cell(row=fila, column=1).value
        if isinstance(val, str) and tipo.lower() in val.lower():
            fila_titulo = fila
            break

    if fila_titulo is None:
        return {"sin_peptido": None, "con_peptido": None, "al_terminar": None}

    fila_datos = fila_titulo + 3

    columnas_por_condicion = {
        "sin_peptido": 1,
        "con_peptido": 7,
        "al_terminar": 13,
    }

    resultado = {}
    for clave, col_base in columnas_por_condicion.items():
        longitud_test = ws.cell(row=fila_datos, column=col_base).value
        if not isinstance(longitud_test, (int, float)):
            resultado[clave] = None
            continue

        longitudes = []
        cubetas = {"Sample": [], "Reference": [], "Blank": [], "Sample2": []}
        orden_cubetas = ["Sample", "Reference", "Blank", "Sample2"]

        fila = fila_datos
        while fila <= ws.max_row:
            longitud = ws.cell(row=fila, column=col_base).value
            if not isinstance(longitud, (int, float)):
                break
            longitudes.append(longitud)
            for i, cub in enumerate(orden_cubetas):
                v = ws.cell(row=fila, column=col_base + 1 + i).value
                cubetas[cub].append(v if isinstance(v, (int, float)) else None)
            fila += 1

        cubetas_validas = {}
        for cub, vals in cubetas.items():
            cubetas_validas[cub] = None if all(v is None for v in vals) else vals

        resultado[clave] = {"longitudes": longitudes, **cubetas_validas}

    return resultado



def _ylim_con_margen(valores_min, valores_max, margen_frac, margen_abs_default=0.01):
    if not valores_min:
        return None
    ymin = min(valores_min)
    ymax = max(valores_max)
    rango = ymax - ymin
    margen = rango * margen_frac if rango > 0 else margen_abs_default
    return (ymin - margen, ymax + margen)


def calcular_ylim_gp_vs_t(cepas, obtener_ws_func, margen_frac=MARGEN_YLIM):
   
    valores_min, valores_max = [], []
    for cepa in cepas:
        for volumen in VOLUMEN_A_PL:
            ws = obtener_ws_func(cepa, volumen)
            if ws is None:
                continue
            temps, gp, desv = leer_bloque_gp(ws)
            if not temps:
                continue
            for g, d in zip(gp, desv):
                valores_min.append(g - d)
                valores_max.append(g + d)
    return _ylim_con_margen(valores_min, valores_max, margen_frac)


def calcular_ylim_intensidad(ws_a, ws_b, tipo, condicion, margen_frac=MARGEN_YLIM):
    
    valores = []
    for ws in (ws_a, ws_b):
        if ws is None:
            continue
        datos = leer_bloque_intensidad(ws, tipo)
        bloque = datos.get(condicion)
        if bloque is None:
            continue
        for cub in ["Sample", "Reference", "Blank", "Sample2"]:
            vals = bloque.get(cub)
            if vals is None:
                continue
            valores.extend(v for v in vals if v is not None)
    if not valores:
        return None
    return _ylim_con_margen(valores, valores, margen_frac)



def grafica_gp_vs_temperatura(wb, cepa, peptido_corto, peptido_bonito,
                               carpeta_salida, ylim=None):
    
    fig, ax = plt.subplots(figsize=(10, 6))
    hay_datos = False

    for volumen, pl in VOLUMEN_A_PL.items():
        ws = buscar_hoja_piam(wb, cepa, peptido_corto, volumen)
        if ws is None:
            continue
        temps, gp, desv = leer_bloque_gp(ws)
        if not temps:
            continue
        ax.errorbar(temps, gp, yerr=desv, marker="o", capsize=2,
                    color=COLOR_PL[pl], label=f"P/L={pl}", markersize=4)
        hay_datos = True

    if not hay_datos:
        plt.close(fig)
        return

    ax.set_xlabel("Temperatura (°C)")
    ax.set_ylabel("GP Laurdan")
    ax.set_title(f"S aureus {cepa} con {peptido_bonito}")
    if ylim is not None:
        ax.set_ylim(*ylim)
    ax.legend(loc="best")
    ax.grid(True, alpha=0.3)

    nombre = f"GP_vs_T_S_aureus_{cepa}_con_{peptido_bonito}.png"
    fig.tight_layout()
    fig.savefig(os.path.join(carpeta_salida, nombre), dpi=150)
    plt.close(fig)
    print(f"  Guardado: {nombre}")


def grafica_gp_vs_temperatura_ll37(wb, cepa, carpeta_salida, ylim=None):
    fig, ax = plt.subplots(figsize=(10, 6))
    hay_datos = False

    for volumen, pl in VOLUMEN_A_PL.items():
        ws = buscar_hoja_ll37(wb, cepa, volumen)
        if ws is None:
            continue
        temps, gp, desv = leer_bloque_gp(ws)
        if not temps:
            continue
        ax.errorbar(temps, gp, yerr=desv, marker="o", capsize=2,
                    color=COLOR_PL[pl], label=f"P/L={pl}", markersize=4)
        hay_datos = True

    if not hay_datos:
        plt.close(fig)
        return

    ax.set_xlabel("Temperatura (°C)")
    ax.set_ylabel("GP Laurdan")
    ax.set_title(f"S aureus {cepa} con LL-37")
    if ylim is not None:
        ax.set_ylim(*ylim)
    ax.legend(loc="best")
    ax.grid(True, alpha=0.3)

    nombre = f"GP_vs_T_S_aureus_{cepa}_con_LL-37.png"
    fig.tight_layout()
    fig.savefig(os.path.join(carpeta_salida, nombre), dpi=150)
    plt.close(fig)
    print(f"  Guardado: {nombre}")


def grafica_gp_vs_t_control_combinada(wb, carpeta_salida):
    
    fig, ax = plt.subplots(figsize=(10, 6))
    hay_datos = False

    for cepa in CEPAS:
        ws = buscar_hoja_piam(wb, cepa, "", "0ul")
        if ws is None:
            continue
        temps, gp, desv = leer_bloque_gp(ws)
        if not temps:
            continue
        ax.errorbar(temps, gp, yerr=desv, marker=MARKER_CEPA[cepa],
                    linestyle="-", color=COLOR_CEPA[cepa], capsize=2,
                    markersize=6, linewidth=1.5, label=f"S aureus {cepa}")
        hay_datos = True

    if not hay_datos:
        plt.close(fig)
        return

    ax.set_xlabel("Temperatura (°C)")
    ax.set_ylabel("GP Laurdan")
    ax.set_title("S aureus 144 y 145 sin péptido (control)")
    ax.legend(loc="best")
    ax.grid(True, alpha=0.3)

    nombre = "GP_vs_T_S_aureus_144_145_control.png"
    fig.tight_layout()
    fig.savefig(os.path.join(carpeta_salida, nombre), dpi=150)
    plt.close(fig)
    print(f"  Guardado: {nombre}")


def grafica_gp_vs_pl_37_combinada(wb, peptido_corto, peptido_bonito,
                                   carpeta_salida, es_ll37=False):
    
    fig, ax = plt.subplots(figsize=(9, 6))
    hay_datos = False

    for cepa in CEPAS:
        pls, gps, desvs = [], [], []
        for volumen, pl in VOLUMEN_A_PL.items():
            ws = obtener_hoja(wb, cepa, peptido_corto, volumen, es_ll37)
            if ws is None:
                continue
            temps, gp, desv = leer_bloque_gp(ws)
            if not temps:
                continue
            try:
                idx = temps.index(37)
            except ValueError:
                idx = min(range(len(temps)), key=lambda i: abs(temps[i] - 37))
            pls.append(pl)
            gps.append(gp[idx])
            desvs.append(desv[idx])

        if not pls:
            continue

        
        orden = sorted(range(len(pls)), key=lambda i: pls[i])
        pls = [pls[i] for i in orden]
        gps = [gps[i] for i in orden]
        desvs = [desvs[i] for i in orden]

        ax.errorbar(pls, gps, yerr=desvs, marker=MARKER_CEPA[cepa],
                    linestyle="-", color=COLOR_CEPA[cepa], capsize=4,
                    markersize=9, linewidth=1.8, label=f"S aureus {cepa}")
        hay_datos = True

    if not hay_datos:
        plt.close(fig)
        return

    ax.set_xlabel("P/L")
    ax.set_ylabel("GP Laurdan")
    ax.set_title(f"GP a 37°C vs P/L con {peptido_bonito}")
    ax.legend(loc="best")
    ax.grid(True, alpha=0.3)

    nombre = f"GP_vs_PL_37C_S_aureus_144_145_con_{peptido_bonito}.png"
    fig.tight_layout()
    fig.savefig(os.path.join(carpeta_salida, nombre), dpi=150)
    plt.close(fig)
    print(f"  Guardado: {nombre}")


def hill_decay(L, y_min, y_max, Kd, n):
   
    return y_min + (y_max - y_min) * (Kd**n / (Kd**n + L**n))


def _datos_y_ajuste_hill(wb, peptido_corto, cepa, es_ll37=False):
    
    pls, gps, desvs = [], [], []
    for volumen, pl in VOLUMEN_A_PL.items():
        ws = obtener_hoja(wb, cepa, peptido_corto, volumen, es_ll37)
        if ws is None:
            continue
        temps, gp, desv = leer_bloque_gp(ws)
        if not temps:
            continue
        try:
            idx = temps.index(37)
        except ValueError:
            idx = min(range(len(temps)), key=lambda i: abs(temps[i] - 37))
        pls.append(pl)
        gps.append(gp[idx])
        desvs.append(desv[idx])

    if len(pls) < 4:
        return {"error": "pocos_datos", "n_puntos": len(pls)}

    orden = sorted(range(len(pls)), key=lambda i: pls[i])
    pls = np.array([pls[i] for i in orden], dtype=float)
    gps = np.array([gps[i] for i in orden], dtype=float)
    desvs = np.array([desvs[i] for i in orden], dtype=float)

    
    gp_base = gps[0]
    gps_trasladado = gps - gp_base

    pl_max = max(pls)

 #ajuste rango kd y n
    Kd_min, Kd_max = 1e-4, 10 * pl_max
    n_min, n_max = 0.2, 8.0

    Kd0 = min(max(pl_max / 2, Kd_min), Kd_max)
    p0 = [min(gps_trasladado), max(gps_trasladado), Kd0, 1.0]
    bounds = ([-np.inf, -np.inf, Kd_min, n_min], [np.inf, np.inf, Kd_max, n_max])

    desvs_fit = np.where(desvs <= 0, 1e-6, desvs) 

    try:
        params, _ = curve_fit(hill_decay, pls, gps_trasladado, sigma=desvs_fit,
                               absolute_sigma=True, p0=p0, bounds=bounds,
                               maxfev=20000)
    except RuntimeError as e:
        return {"error": "no_convergio", "detalle": str(e)}

    return {
        "error": None,
        "pls": pls, "gps": gps, "desvs": desvs,
        "gps_trasladado": gps_trasladado, "gp_base": gp_base,
        "params": params,
        "Kd_min": Kd_min, "Kd_max": Kd_max,
        "n_min": n_min, "n_max": n_max,
    }


def calcular_lims_hill(wb, peptido_corto, cepas, es_ll37=False,
                        margen_frac=MARGEN_YLIM):
   
    y_min_vals, y_max_vals = [], []
    x_min_vals, x_max_vals = [], []

    for cepa in cepas:
        info = _datos_y_ajuste_hill(wb, peptido_corto, cepa, es_ll37=es_ll37)
        if info.get("error") is not None:
            continue

        pls = info["pls"]
        gps_trasladado = info["gps_trasladado"]
        desvs = info["desvs"]

        for g, d in zip(gps_trasladado, desvs):
            y_min_vals.append(g - d)
            y_max_vals.append(g + d)

        L_fit = np.linspace(0, max(pls) * 1.1, 300)
        GP_fit = hill_decay(L_fit, *info["params"])
        y_min_vals.append(float(np.min(GP_fit)))
        y_max_vals.append(float(np.max(GP_fit)))

        x_min_vals.append(float(np.min(L_fit)))
        x_max_vals.append(float(np.max(L_fit)))

    ylim = _ylim_con_margen(y_min_vals, y_max_vals, margen_frac)
    xlim = _ylim_con_margen(x_min_vals, x_max_vals, margen_frac)
    return ylim, xlim


def ajustar_hill_gp_vs_pl37(wb, peptido_corto, peptido_bonito, cepa,
                             carpeta_salida, es_ll37=False,
                             ylim=None, xlim=None):
    
    info = _datos_y_ajuste_hill(wb, peptido_corto, cepa, es_ll37=es_ll37)

    if info.get("error") == "pocos_datos":
        print(f"  [Hill] {cepa} {peptido_bonito}: datos insuficientes "
              f"({info['n_puntos']} puntos), se omite")
        return None
    if info.get("error") == "no_convergio":
        print(f"  [Hill] {cepa} {peptido_bonito}: no convergio ({info['detalle']})")
        return None

    pls = info["pls"]
    desvs = info["desvs"]
    gps_trasladado = info["gps_trasladado"]
    gp_base = info["gp_base"]
    params = info["params"]
    Kd_min, Kd_max = info["Kd_min"], info["Kd_max"]
    n_min, n_max = info["n_min"], info["n_max"]

    y_min_fit, y_max_fit, Kd_fit, n_fit = params

   
    avisos = []
    if Kd_fit >= Kd_max * 0.99:
        avisos.append("el efecto no muestra saturación clara en el rango de P/L medido")
    if Kd_fit <= Kd_min * 1.01:
        avisos.append("el efecto parece saturar antes de la dosis más baja medida")
    if n_fit >= n_max * 0.99:
        avisos.append("transición muy abrupta: n podría ser mayor a lo mostrado")
    if n_fit <= n_min * 1.01:
        avisos.append("transición muy gradual: n podría ser menor a lo mostrado")

    pegado_a_limite = len(avisos) > 0
    if pegado_a_limite:
        print(f"  [Hill] AVISO (solo consola, no se muestra en la grafica) "
              f"{cepa} {peptido_bonito}: " + "; ".join(avisos) +
              f"  (Kd={Kd_fit:.4g}, n={n_fit:.4g}). Con solo 4 puntos, tomar "
              f"el(los) parametro(s) senalado(s) solo como referencia.")

    L_fit = np.linspace(0, max(pls) * 1.1, 300)
    GP_fit = hill_decay(L_fit, *params)

    fig, ax = plt.subplots(figsize=(9, 7))
    ax.errorbar(pls, gps_trasladado, yerr=desvs, fmt='o', capsize=5,
                markersize=8, color='tab:blue', label='Datos')
    ax.plot(L_fit, GP_fit, color='tab:red', linewidth=2, label='Ajuste de Hill')

    ax.set_xlabel('P/L')
    ax.set_ylabel('GP Laurdan')
    ax.set_title(f'Ajuste de Hill - S aureus {cepa} con {peptido_bonito} a 37°C')
    if ylim is not None:
        ax.set_ylim(*ylim)
    if xlim is not None:
        ax.set_xlim(*xlim)
    ax.legend(loc='lower right')
    ax.grid(True, alpha=0.3)

    
    if n_fit >= n_max * 0.99:
        texto_n = f"$n$ > {n_max:.4g}"
    elif n_fit <= n_min * 1.01:
        texto_n = f"$n$ < {n_min:.4g}"
    else:
        texto_n = f"$n$ = {n_fit:.4g}"

    texto_parametros = (
        f"$Y_{{max}}$ = {y_min_fit:.4g}\n"
        f"$K_d$ = {Kd_fit:.4g}\n"
        f"{texto_n}"
    )
    color_caja = 'mistyrose' if pegado_a_limite else 'white'
    color_texto = 'firebrick' if pegado_a_limite else 'black'
    ax.text(0.05, 0.95, texto_parametros, transform=ax.transAxes,
            fontsize=15, verticalalignment='top', color=color_texto,
            bbox=dict(boxstyle='round', facecolor=color_caja, alpha=0.9,
                       edgecolor='gray'))

    nombre = f"Hill_GP_vs_PL_37C_S_aureus_{cepa}_con_{peptido_bonito}.png"
    fig.tight_layout()
    fig.savefig(os.path.join(carpeta_salida, nombre), dpi=150)
    plt.close(fig)
    print(f"  Guardado: {nombre}  "
          f"(y_min={y_min_fit:.4f}, y_max={y_max_fit:.4f}, "
          f"Kd={Kd_fit:.4f}, n={n_fit:.4f})")

    return {
        "cepa": cepa, "peptido": peptido_bonito,
        "gp_base_P_L_0": gp_base,
        "y_min": y_min_fit, "y_max": y_max_fit,
        "Kd": Kd_fit, "n": n_fit,
    }


def etiqueta_caso(cepa, peptido_bonito, volumen):
    
    if volumen == "0ul":
        return f"S aureus {cepa} control (P/L=0)"
    pl = VOLUMEN_A_PL[volumen]
    return f"S aureus {cepa} con {peptido_bonito} (P/L={pl})"


def grafica_intensidad_por_condicion(wb, nombre_hoja, tipo, carpeta_destino,
                                      cepa, peptido_bonito, volumen,
                                      ylim_por_condicion=None):
   
    ws = wb[nombre_hoja]
    datos = leer_bloque_intensidad(ws, tipo)

    if ylim_por_condicion is None:
        ylim_por_condicion = {}

    etiquetas_bonitas = {
        "sin_peptido": "sin peptido",
        "con_peptido": "con peptido",
        "al_terminar": "con peptido al terminar",
    }
    colores_cubeta = {
        "Sample":    "tab:blue",
        "Reference": "tab:red",
        "Blank":     "tab:green",
        "Sample2":   "tab:cyan",
    }

    for clave, etiqueta in etiquetas_bonitas.items():
        bloque = datos.get(clave)
        if bloque is None:
            continue

        fig, ax = plt.subplots(figsize=(9, 6))
        longitudes = bloque["longitudes"]
        hay_alguna_cubeta = False

        for cub in ["Sample", "Reference", "Blank", "Sample2"]:
            vals = bloque.get(cub)
            if vals is None:
                continue
            x = [l for l, v in zip(longitudes, vals) if v is not None]
            y = [v for v in vals if v is not None]
            if not y:
                continue
            ax.plot(x, y, color=colores_cubeta[cub], label=cub, linewidth=1.5)
            hay_alguna_cubeta = True

        if not hay_alguna_cubeta:
            plt.close(fig)
            continue

        ax.set_xlabel("Longitud de onda (nm)")
        ax.set_ylabel("Intensidad")
        
        if clave == "sin_peptido":
            titulo = f"S aureus {cepa}"
        else:
            titulo = etiqueta_caso(cepa, peptido_bonito, volumen)
        ax.set_title(titulo)
        if clave in ylim_por_condicion:
            ax.set_ylim(*ylim_por_condicion[clave])
        ax.legend(loc="best")
        ax.grid(True, alpha=0.3)

        etiqueta_safe = etiqueta.replace(" ", "_")
        nombre = f"{tipo}_{etiqueta_safe}.png"
        fig.tight_layout()
        fig.savefig(os.path.join(carpeta_destino, nombre), dpi=150)
        plt.close(fig)
        print(f"    Guardado: {nombre}")


def procesar_espectro_pareja(wb, nombre_hoja_144, nombre_hoja_145,
                              subcarpeta_144, subcarpeta_145, carpeta_salida,
                              peptido_bonito, volumen):
   
    if nombre_hoja_144 is None and nombre_hoja_145 is None:
        return

    tipo = "Espectro"
    ws_144 = wb[nombre_hoja_144] if nombre_hoja_144 else None
    ws_145 = wb[nombre_hoja_145] if nombre_hoja_145 else None

    ylim_por_condicion = {}
    for condicion in ["sin_peptido", "con_peptido", "al_terminar"]:
        ylim = calcular_ylim_intensidad(ws_144, ws_145, tipo, condicion)
        if ylim is not None:
            ylim_por_condicion[condicion] = ylim

    if nombre_hoja_144:
        carpeta_144 = os.path.join(carpeta_salida, subcarpeta_144)
        os.makedirs(carpeta_144, exist_ok=True)
        grafica_intensidad_por_condicion(wb, nombre_hoja_144, tipo,
                                          carpeta_144, cepa="144",
                                          peptido_bonito=peptido_bonito,
                                          volumen=volumen,
                                          ylim_por_condicion=ylim_por_condicion)
    if nombre_hoja_145:
        carpeta_145 = os.path.join(carpeta_salida, subcarpeta_145)
        os.makedirs(carpeta_145, exist_ok=True)
        grafica_intensidad_por_condicion(wb, nombre_hoja_145, tipo,
                                          carpeta_145, cepa="145",
                                          peptido_bonito=peptido_bonito,
                                          volumen=volumen,
                                          ylim_por_condicion=ylim_por_condicion)


def maximos_por_cubeta(bloque):
   
    if bloque is None:
        return []
    maximos = []
    for cub in ["Sample", "Reference", "Blank", "Sample2"]:
        vals = bloque.get(cub)
        if vals is None:
            continue
        vals_validos = [v for v in vals if v is not None]
        if not vals_validos:
            continue
        maximos.append(max(vals_validos))
    return maximos


def nivel_maximo_promedio(bloque):
    
    maximos = maximos_por_cubeta(bloque)
    if not maximos:
        return None, None
    promedio = sum(maximos) / len(maximos)
    if len(maximos) > 1:
        desviacion = float(np.std(maximos, ddof=1))
    else:
        desviacion = 0.0
    return promedio, desviacion


def p_a_asteriscos(p):
   
    if p < 0.0001:
        return "****"
    if p < 0.001:
        return "***"
    if p < 0.01:
        return "**"
    if p < 0.05:
        return "*"
    return "ns"


def agregar_significancia(ax, x1, x2, y, texto, color="black"):
    
    ymin, ymax = ax.get_ylim()
    h = (ymax - ymin) * 0.02
    ax.plot([x1, x1, x2, x2], [y, y + h, y + h, y], lw=1.3, color=color)
    ax.text((x1 + x2) / 2, y + h, texto, ha="center", va="bottom",
             fontsize=14, color=color)


def grafica_histograma_dispersion(wb, nombre_hoja, carpeta_destino,
                                   cepa, peptido_bonito, volumen):
   
    ws = wb[nombre_hoja]
    datos = leer_bloque_intensidad(ws, "Dispersion")

    nivel_sin, desv_sin = nivel_maximo_promedio(datos.get("sin_peptido"))
    nivel_con, desv_con = nivel_maximo_promedio(datos.get("con_peptido"))

   
    p_valor, tipo_prueba = prueba_estadistica_sin_vs_con(
        datos.get("sin_peptido"), datos.get("con_peptido"))
    if p_valor is not None:
        print(f"    [Dispersion] {etiqueta_caso(cepa, peptido_bonito, volumen)}: "
              f"prueba t {tipo_prueba}, p={p_valor:.4g} ({p_a_asteriscos(p_valor)})")

    
    tiempos_reales, niveles, errores, colores, etiquetas = [], [], [], [], []
    if nivel_sin is not None:
        tiempos_reales.append(0)
        niveles.append(nivel_sin)
        errores.append(desv_sin)
        colores.append("tab:blue")
        etiquetas.append("0\n(sin péptido)")
    if nivel_con is not None:
        tiempos_reales.append(TIEMPO_CON_PEPTIDO_MIN)
        niveles.append(nivel_con)
        errores.append(desv_con)
        colores.append("tab:orange")
        etiquetas.append(f"{TIEMPO_CON_PEPTIDO_MIN}\n(con péptido)")

    if not niveles:
        return

   
    posiciones = np.arange(len(niveles))
    ancho_barra = 0.6
    margen_x = 0.45

    fig, ax = plt.subplots(figsize=(5, 6))
    ax.bar(posiciones, niveles, yerr=errores, color=colores, width=ancho_barra,
           capsize=6, error_kw=dict(ecolor="black", elinewidth=1.3))
    ax.set_xticks(posiciones)
    ax.set_xticklabels(etiquetas)
    ax.set_xlabel("Tiempo (min)")
    ax.set_ylabel("U.A.")
    ax.set_title(etiqueta_caso(cepa, peptido_bonito, volumen), fontsize=11)
    ax.grid(True, axis="y", alpha=0.3)

    
    if len(niveles) == 2 and p_valor is not None and p_valor < 0.05:
        ymax_datos = max(n + e for n, e in zip(niveles, errores))
        ax.set_ylim(0, ymax_datos * 1.28)
        agregar_significancia(ax, posiciones[0], posiciones[1],
                               ymax_datos * 1.08, p_a_asteriscos(p_valor))

    ax.set_xlim(posiciones[0] - margen_x, posiciones[-1] + margen_x)

    nombre = "Dispersion_histograma.png"
    fig.tight_layout()
    fig.savefig(os.path.join(carpeta_destino, nombre), dpi=150)
    plt.close(fig)
    print(f"    Guardado: {nombre}  "
          f"(sin={nivel_sin}±{desv_sin}, con={nivel_con}±{desv_con})")


def prueba_estadistica_sin_vs_con(bloque_sin, bloque_con):
   
    if bloque_sin is None or bloque_con is None:
        return None, None

    pares_sin, pares_con = [], []
    for cub in ["Sample", "Reference", "Blank", "Sample2"]:
        vals_sin = bloque_sin.get(cub)
        vals_con = bloque_con.get(cub)
        if vals_sin is None or vals_con is None:
            continue
        validos_sin = [v for v in vals_sin if v is not None]
        validos_con = [v for v in vals_con if v is not None]
        if not validos_sin or not validos_con:
            continue
        pares_sin.append(max(validos_sin))
        pares_con.append(max(validos_con))

    if len(pares_sin) >= 2:
        try:
            resultado = ttest_rel(pares_con, pares_sin)
            return float(resultado.pvalue), "pareada"
        except Exception:
            pass

  
    vals_sin = maximos_por_cubeta(bloque_sin)
    vals_con = maximos_por_cubeta(bloque_con)
    if len(vals_sin) >= 2 and len(vals_con) >= 2:
        try:
            resultado = ttest_ind(vals_con, vals_sin)
            return float(resultado.pvalue), "no pareada"
        except Exception:
            pass

    return None, None


def limpiar_carpetas_vacias(carpeta_raiz):
    
    eliminadas = []
    for carpeta_actual, _subcarpetas, _archivos in os.walk(carpeta_raiz, topdown=False):
        if carpeta_actual == carpeta_raiz:
            continue  
        if not os.listdir(carpeta_actual):
            os.rmdir(carpeta_actual)
            eliminadas.append(carpeta_actual)
    if eliminadas:
        print(f"\n  Carpetas vacias eliminadas ({len(eliminadas)}):")
        for c in eliminadas:
            print(f"    - {os.path.relpath(c, carpeta_raiz)}")


def main(archivo_excel):
    if not os.path.isfile(archivo_excel):
        print(f"No se encuentra el archivo: {archivo_excel}")
        return

    carpeta_salida = os.path.join(
        os.path.dirname(os.path.abspath(archivo_excel)), "graficas"
    )
   
    if os.path.isdir(carpeta_salida):
        shutil.rmtree(carpeta_salida)
    os.makedirs(carpeta_salida, exist_ok=True)

    print(f"Cargando: {archivo_excel}")
    wb = cargar_excel(archivo_excel)

    
    print("\n--- GP vs T (piam) -- y1/y2 compartidos por pareja 144/145 ---")
    for peptido_corto, peptido_bonito in PEPTIDOS_PIAM.items():
        ylim = calcular_ylim_gp_vs_t(
            CEPAS,
            lambda cepa, volumen, pc=peptido_corto: buscar_hoja_piam(wb, cepa, pc, volumen),
        )
        print(f"  {peptido_bonito}: ylim compartido = {ylim}")
        for cepa in CEPAS:
            grafica_gp_vs_temperatura(wb, cepa, peptido_corto, peptido_bonito,
                                       carpeta_salida, ylim=ylim)

    print("\n--- GP vs T (LL-37) -- y1/y2 compartidos por pareja 144/145 ---")
    ylim_ll37 = calcular_ylim_gp_vs_t(
        CEPAS, lambda cepa, volumen: buscar_hoja_ll37(wb, cepa, volumen),
    )
    print(f"  LL-37: ylim compartido = {ylim_ll37}")
    for cepa in CEPAS:
        grafica_gp_vs_temperatura_ll37(wb, cepa, carpeta_salida, ylim=ylim_ll37)

    print("\n--- GP vs T control (sin péptido) -- 144 y 145 superpuestos ---")
    grafica_gp_vs_t_control_combinada(wb, carpeta_salida)

  
    print("\n--- GP vs P/L a 37C -- 144 y 145 superpuestos en una sola grafica ---")
    for peptido_corto, peptido_bonito in PEPTIDOS.items():
        es_ll37 = (peptido_corto == "LL-37")
        grafica_gp_vs_pl_37_combinada(wb, peptido_corto, peptido_bonito,
                                       carpeta_salida, es_ll37=es_ll37)

    
    print("\n--- Ajuste de Hill (GP vs P/L a 37C), individual por cepa "
          "-- ejes compartidos por pareja 144/145 ---")
    for peptido_corto, peptido_bonito in PEPTIDOS.items():
        es_ll37 = (peptido_corto == "LL-37")
        ylim_hill, xlim_hill = calcular_lims_hill(wb, peptido_corto, CEPAS,
                                                   es_ll37=es_ll37)
        print(f"  {peptido_bonito}: ylim compartido = {ylim_hill}, "
              f"xlim compartido = {xlim_hill}")
        for cepa in CEPAS:
            ajustar_hill_gp_vs_pl37(wb, peptido_corto, peptido_bonito,
                                     cepa, carpeta_salida, es_ll37=es_ll37,
                                     ylim=ylim_hill, xlim=xlim_hill)

    
    print("\n--- Control (0uL): Espectro pareado + histograma Dispersion ---")
    ws_144_0 = buscar_hoja_piam(wb, "144", "", "0ul")
    ws_145_0 = buscar_hoja_piam(wb, "145", "", "0ul")
    nombre_144_0 = ws_144_0.title if ws_144_0 else None
    nombre_145_0 = ws_145_0.title if ws_145_0 else None

    procesar_espectro_pareja(wb, nombre_144_0, nombre_145_0,
                              "sa_144-control_0ul", "sa_145-control_0ul",
                              carpeta_salida, peptido_bonito=None, volumen="0ul")

    for nombre_hoja, subcarpeta, cepa_actual in [(nombre_144_0, "sa_144-control_0ul", "144"),
                                                   (nombre_145_0, "sa_145-control_0ul", "145")]:
        if nombre_hoja is None:
            continue
        carpeta = os.path.join(carpeta_salida, subcarpeta)
        os.makedirs(carpeta, exist_ok=True)
        grafica_histograma_dispersion(wb, nombre_hoja, carpeta,
                                       cepa=cepa_actual, peptido_bonito=None, volumen="0ul")

    print("\n--- piam y LL-37: Espectro pareado + histograma Dispersion (solo piam) ---")
    for peptido_corto, peptido_bonito in PEPTIDOS.items():
        es_ll37 = (peptido_corto == "LL-37")
        peptido_safe = peptido_bonito.replace("-", "_")

        for volumen in VOLUMEN_A_PL:
            if volumen == "0ul":
                continue  

            ws_144 = obtener_hoja(wb, "144", peptido_corto, volumen, es_ll37)
            ws_145 = obtener_hoja(wb, "145", peptido_corto, volumen, es_ll37)
            nombre_144 = ws_144.title if ws_144 else None
            nombre_145 = ws_145.title if ws_145 else None

            if nombre_144 is None and nombre_145 is None:
                continue

            subcarpeta_144 = f"sa_144-{peptido_safe}-{volumen}"
            subcarpeta_145 = f"sa_145-{peptido_safe}-{volumen}"
            print(f"  {peptido_bonito} {volumen}: {nombre_144} <-> {nombre_145}")

           
            procesar_espectro_pareja(wb, nombre_144, nombre_145,
                                      subcarpeta_144, subcarpeta_145,
                                      carpeta_salida, peptido_bonito=peptido_bonito,
                                      volumen=volumen)

            
            if not es_ll37:
                if nombre_144:
                    carpeta_144 = os.path.join(carpeta_salida, subcarpeta_144)
                    os.makedirs(carpeta_144, exist_ok=True)
                    grafica_histograma_dispersion(wb, nombre_144, carpeta_144,
                                                   cepa="144", peptido_bonito=peptido_bonito,
                                                   volumen=volumen)
                if nombre_145:
                    carpeta_145 = os.path.join(carpeta_salida, subcarpeta_145)
                    os.makedirs(carpeta_145, exist_ok=True)
                    grafica_histograma_dispersion(wb, nombre_145, carpeta_145,
                                                   cepa="145", peptido_bonito=peptido_bonito,
                                                   volumen=volumen)

    limpiar_carpetas_vacias(carpeta_salida)
    print(f"\nListo. Graficas guardadas en: {carpeta_salida}")


if __name__ == "__main__":
    ARCHIVO_EXCEL = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "Excel de datos.xlsx"
    )
    main(ARCHIVO_EXCEL)
