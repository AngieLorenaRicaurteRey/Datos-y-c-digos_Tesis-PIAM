CARPETAS_EXCLUIR = {
    "sa 144-piam217-2ul dudoso",
}


RENOMBRAR_HOJA = {
    "sa 144-piam217-2ul verificacion": "sa 144-piam217-2ul",
}

import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

def leer_ifx(ruta):
    
    with open(ruta, "r", encoding="latin-1") as f:
        contenido = f.read()

    if "[Data]" not in contenido:
        return []

    bloque_datos = contenido.split("[Data]", 1)[1].strip()
    filas = []
    for linea in bloque_datos.splitlines():
        linea = linea.strip()
        if not linea:
            continue
        partes = [p.strip() for p in linea.split("\t")]
        filas.append(partes)
    return filas


def procesar_gp(ruta_ifx):
   
    filas = leer_ifx(ruta_ifx)

    
    datos = {}
    for fila in filas:
        if len(fila) < 5:
            continue
        try:
            temperatura = float(fila[0])
            cubeta = fila[1]
            gp_valor = float(fila[4])  
        except ValueError:
            continue

        temp_int = int(round(temperatura))
        if temp_int not in datos:
            datos[temp_int] = {}
        datos[temp_int][cubeta] = gp_valor

    if not datos:
        return pd.DataFrame()

    orden_cubetas = ["Sample", "Reference", "Blank", "Sample2"]
    filas_df = []
    for temp in sorted(datos.keys()):
        valores = datos[temp]
        fila_df = {"Temperatura": temp}
        for cubeta in orden_cubetas:
            fila_df[cubeta] = valores.get(cubeta, None)
        filas_df.append(fila_df)

    columnas = ["Temperatura"] + orden_cubetas
    return pd.DataFrame(filas_df, columns=columnas)



def procesar_intensidad_vs_longitud(ruta_ifx):
    
    filas = leer_ifx(ruta_ifx)

    datos = {}
    for fila in filas:
        if len(fila) < 3:
            continue
        try:
            cubeta = fila[0]
            longitud = float(fila[1])
            intensidad = float(fila[2])
        except ValueError:
            continue

        long_int = int(round(longitud))
        if long_int not in datos:
            datos[long_int] = {}
        datos[long_int][cubeta] = intensidad

    if not datos:
        return pd.DataFrame()

    orden_cubetas = ["Sample", "Reference", "Blank", "Sample2"]
    filas_df = []
    for longitud in sorted(datos.keys()):
        valores = datos[longitud]
        fila_df = {"Longitud de onda": longitud}
        for cubeta in orden_cubetas:
            fila_df[cubeta] = valores.get(cubeta, None)
        filas_df.append(fila_df)

    columnas = ["Longitud de onda"] + orden_cubetas
    return pd.DataFrame(filas_df, columns=columnas)



def clasificar_archivos(carpeta):
    
    archivos = {
        "gp": None,
        "dispersion_sin": None,
        "dispersion_con": None,
        "dispersion_final": None,
        "espectro_sin": None,
        "espectro_con": None,
        "espectro_final": None,
    }

    for nombre in os.listdir(carpeta):
        if not nombre.lower().endswith(".ifx"):
            continue
        ruta = os.path.join(carpeta, nombre)
        nombre_lower = nombre.lower()

       
        if "cu de" in nombre_lower:
            archivos["gp"] = ruta
        elif nombre_lower.startswith("dipersion") or nombre_lower.startswith("dispersion"):
            if "al terminar" in nombre_lower:
                archivos["dispersion_final"] = ruta
            elif "piam" in nombre_lower:
                archivos["dispersion_con"] = ruta
            else:
                archivos["dispersion_sin"] = ruta
        elif nombre_lower.startswith("espectro"):
            if "al terminar" in nombre_lower:
                archivos["espectro_final"] = ruta
            elif "piam" in nombre_lower:
                archivos["espectro_con"] = ruta
            else:
                archivos["espectro_sin"] = ruta

    return archivos



def escribir_hoja(wb, nombre_hoja, df_gp, dispersiones, espectros):
    
    nombre_hoja = nombre_hoja[:31]
    ws = wb.create_sheet(nombre_hoja)

    estilo_titulo = Font(bold=True, color="FFFFFF")
    relleno_titulo = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    estilo_subtitulo = Font(bold=True)
    relleno_subtitulo = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    centrado = Alignment(horizontal="center")

    fila_actual = 1


    ws.cell(row=fila_actual, column=1, value="GP vs Temperatura").font = estilo_titulo
    ws.cell(row=fila_actual, column=1).fill = relleno_titulo
    ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=7)
    fila_actual += 1

    if df_gp is not None and not df_gp.empty:
      
        encabezados_gp = list(df_gp.columns) + ["GP promedio", "Desviacion GP"]
        for j, col in enumerate(encabezados_gp, start=1):
            c = ws.cell(row=fila_actual, column=j, value=col)
            c.font = estilo_subtitulo
            c.fill = relleno_subtitulo
            c.alignment = centrado
        fila_actual += 1
        
        for _, row in df_gp.iterrows():
            for j, col in enumerate(df_gp.columns, start=1):
                ws.cell(row=fila_actual, column=j, value=row[col])
            # columna F = AVERAGE(B:E), columna G = STDEV(B:E)
            ws.cell(row=fila_actual, column=6,
                    value=f"=AVERAGE(B{fila_actual}:E{fila_actual})")
            ws.cell(row=fila_actual, column=7,
                    value=f"=STDEV(B{fila_actual}:E{fila_actual})")
            fila_actual += 1
    else:
        ws.cell(row=fila_actual, column=1, value="(sin datos)")
        fila_actual += 1

    fila_actual += 2  

  
    ws.cell(row=fila_actual, column=1, value="Dispersion (Intensidad vs Longitud de onda)").font = estilo_titulo
    ws.cell(row=fila_actual, column=1).fill = relleno_titulo
    ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=15)
    fila_actual += 1

    fila_actual = escribir_bloque_intensidad(
        ws, fila_actual, dispersiones,
        estilo_subtitulo, relleno_subtitulo, centrado
    )

    fila_actual += 2

 
    ws.cell(row=fila_actual, column=1, value="Espectro (Intensidad vs Longitud de onda)").font = estilo_titulo
    ws.cell(row=fila_actual, column=1).fill = relleno_titulo
    ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=15)
    fila_actual += 1

    escribir_bloque_intensidad(
        ws, fila_actual, espectros,
        estilo_subtitulo, relleno_subtitulo, centrado
    )

   
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def escribir_bloque_intensidad(ws, fila_inicio, dfs_dict, estilo_subtitulo,
                                relleno_subtitulo, centrado):
   
    etiquetas = [
        ("sin_peptido", "Sin peptido"),
        ("con_peptido", "Con peptido"),
        ("al_terminar", "Con peptido al terminar"),
    ]

   
    col_inicio = 1
    for clave, etiqueta in etiquetas:
        c = ws.cell(row=fila_inicio, column=col_inicio, value=etiqueta)
        c.font = estilo_subtitulo
        c.fill = relleno_subtitulo
        c.alignment = centrado
        ws.merge_cells(start_row=fila_inicio, start_column=col_inicio,
                       end_row=fila_inicio, end_column=col_inicio + 4)
        col_inicio += 6  

    fila_actual = fila_inicio + 1

  
    col_inicio = 1
    for clave, _ in etiquetas:
        encabezados = ["Longitud de onda", "Sample", "Reference", "Blank", "Sample2"]
        for j, enc in enumerate(encabezados):
            c = ws.cell(row=fila_actual, column=col_inicio + j, value=enc)
            c.font = estilo_subtitulo
            c.alignment = centrado
        col_inicio += 6
    fila_actual += 1

   
    max_filas = 0
    for clave, _ in etiquetas:
        df = dfs_dict.get(clave)
        if df is not None and not df.empty:
            max_filas = max(max_filas, len(df))

    for i in range(max_filas):
        col_inicio = 1
        for clave, _ in etiquetas:
            df = dfs_dict.get(clave)
            if df is not None and not df.empty and i < len(df):
                fila_datos = df.iloc[i]
                ws.cell(row=fila_actual + i, column=col_inicio,
                        value=fila_datos["Longitud de onda"])
                ws.cell(row=fila_actual + i, column=col_inicio + 1,
                        value=fila_datos.get("Sample"))
                ws.cell(row=fila_actual + i, column=col_inicio + 2,
                        value=fila_datos.get("Reference"))
                ws.cell(row=fila_actual + i, column=col_inicio + 3,
                        value=fila_datos.get("Blank"))
                ws.cell(row=fila_actual + i, column=col_inicio + 4,
                        value=fila_datos.get("Sample2"))
            col_inicio += 6

    return fila_actual + max_filas



def procesar_carpeta(ruta_carpeta):
    
    archivos = clasificar_archivos(ruta_carpeta)

    df_gp = procesar_gp(archivos["gp"]) if archivos["gp"] else None

    dispersiones = {
        "sin_peptido": procesar_intensidad_vs_longitud(archivos["dispersion_sin"])
                        if archivos["dispersion_sin"] else None,
        "con_peptido": procesar_intensidad_vs_longitud(archivos["dispersion_con"])
                        if archivos["dispersion_con"] else None,
        "al_terminar": procesar_intensidad_vs_longitud(archivos["dispersion_final"])
                        if archivos["dispersion_final"] else None,
    }

    espectros = {
        "sin_peptido": procesar_intensidad_vs_longitud(archivos["espectro_sin"])
                        if archivos["espectro_sin"] else None,
        "con_peptido": procesar_intensidad_vs_longitud(archivos["espectro_con"])
                        if archivos["espectro_con"] else None,
        "al_terminar": procesar_intensidad_vs_longitud(archivos["espectro_final"])
                        if archivos["espectro_final"] else None,
    }

    return df_gp, dispersiones, espectros



def main():
    
    ruta_base = os.path.dirname(os.path.abspath(__file__))

  
    subcarpetas = sorted([
        nombre for nombre in os.listdir(ruta_base)
        if os.path.isdir(os.path.join(ruta_base, nombre))
    ])

    if not subcarpetas:
        print("No se encontraron subcarpetas en", ruta_base)
        return

   
    wb = Workbook()
   
    wb.remove(wb.active)

    for nombre_carpeta in subcarpetas:
        
        if nombre_carpeta in CARPETAS_EXCLUIR:
            print(f"Saltando (excluida): {nombre_carpeta}")
            continue

        ruta_carpeta = os.path.join(ruta_base, nombre_carpeta)
     
        nombre_hoja = RENOMBRAR_HOJA.get(nombre_carpeta, nombre_carpeta)
        print(f"Procesando: {nombre_carpeta} -> hoja '{nombre_hoja}'")

        try:
            df_gp, dispersiones, espectros = procesar_carpeta(ruta_carpeta)
            escribir_hoja(wb, nombre_hoja, df_gp, dispersiones, espectros)
        except Exception as e:
            print(f"  Error procesando {nombre_carpeta}: {e}")


    ruta_salida = os.path.join(ruta_base, "resultados.xlsx")
    wb.save(ruta_salida)
    print(f"\nListo. Archivo guardado en: {ruta_salida}")


if __name__ == "__main__":
    main()

